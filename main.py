# Copyright (c) 2025 Alan He. Licensed under MIT.

import argparse
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from datetime import date
from modeling.data import get_historical_financials, get_company_share_float, fetch_company_profile, fetch_market_risk_premium, format_summary_df
from modeling.dcf import calculate_dcf, print_dcf_results, sensitivity_analysis, wacc_sensitivity_analysis, calculate_wacc, print_wacc_details, get_risk_free_rate
from modeling.constants import HISTORICAL_DATA_PERIODS_ANNUAL, HISTORICAL_DATA_PERIODS_QUARTER, TERMINAL_RISK_PREMIUM, TERMINAL_RONIC_PREMIUM
from modeling.ai_analyst import analyze_company, interactive_review, analyze_valuation_gap

current_dir = os.path.dirname(os.path.abspath(__file__))
EXCEL_TEMPLATE_PATH = os.path.join(current_dir, 'modeling', 'DCF valuation template.xlsx')
EXCEL_OUTPUT_DIR = os.path.join(os.getcwd(), 'stock_valuation')

def write_to_excel(filename, base_year_data, financial_data, valuation_params, company_profile, total_equity_risk_premium, gap_analysis_result=None):
    if not os.path.exists(EXCEL_OUTPUT_DIR):
        os.makedirs(EXCEL_OUTPUT_DIR)
    shutil.copy(EXCEL_TEMPLATE_PATH, filename)
    wb = load_workbook(filename)
    ws1 = wb['Input and sensitivity']
    ws2 = wb.create_sheet('Historical Financial Data')
    ws3 = wb.create_sheet('Income Statement')
    ws4 = wb.create_sheet('Balance Sheet')
    ws5 = wb.create_sheet('Cash Flow Statement')
    ws6 = wb['Valuation output']

    ws1.cell(row=2, column=2).value = valuation_params['base_year']
    ws1.cell(row=3, column=2).value = valuation_params['revenue_growth_1'] / 100
    ws1.cell(row=4, column=2).value = valuation_params['revenue_growth_2'] / 100
    ws1.cell(row=5, column=2).value = valuation_params['risk_free_rate']
    ws1.cell(row=6, column=2).value = valuation_params['ebit_margin'] / 100
    ws1.cell(row=7, column=2).value = valuation_params['convergence']
    ws1.cell(row=8, column=2).value = valuation_params['revenue_invested_capital_ratio_1']
    ws1.cell(row=9, column=2).value = valuation_params['revenue_invested_capital_ratio_2']
    ws1.cell(row=10, column=2).value = valuation_params['revenue_invested_capital_ratio_3']
    ws1.cell(row=11, column=2).value = valuation_params['wacc'] / 100
    ws1.cell(row=12, column=2).value = get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM
    ws1.cell(row=13, column=2).value = valuation_params['ronic']
    ws1.cell(row=14, column=2).value = valuation_params['tax_rate'] / 100

    # Update sensitivity table base values (E8 = revenue_growth_2, K2 = ebit_margin)
    ws1.cell(row=8, column=5).value = valuation_params['revenue_growth_2'] / 100   # E8
    ws1.cell(row=2, column=11).value = valuation_params['ebit_margin'] / 100        # K2

    ws1.cell(row=17, column=2).value = valuation_params['risk_free_rate']
    ws1.cell(row=18, column=2).value = base_year_data.get('Cost of Debt', 0) / 100
    ws1.cell(row=19, column=2).value = total_equity_risk_premium
    ws1.cell(row=20, column=2).value = company_profile.get('beta', 1.0)

    for r in dataframe_to_rows(financial_data['summary'], index=True, header=True):
        ws2.append(r)

    for r in dataframe_to_rows(financial_data['income_statement'], index=True, header=True):
        ws3.append(r)

    for r in dataframe_to_rows(financial_data['balance_sheet'], index=True, header=True):
        ws4.append(r)

    for r in dataframe_to_rows(financial_data['cashflow_statement'], index=True, header=True):
        ws5.append(r)

    company_name = company_profile.get('companyName', 'N/A')
    ws6.cell(row=1, column=1).value = f"{company_name} - in {base_year_data.get('Reported Currency', '')}, millions"

    ws6.cell(row=3, column=2).value = float(base_year_data['Revenue Growth']) / 100
    ws6.cell(row=4, column=2).value = float(base_year_data['Revenue'])
    ws6.cell(row=6, column=2).value = float(base_year_data['EBIT'])
    ws6.cell(row=9, column=2).value = float(base_year_data['Total Reinvestments'])
    ws6.cell(row=22, column=2).value = float(base_year_data['Cash & Cash Equivalents'])
    ws6.cell(row=23, column=2).value = float(base_year_data['Total Investments'])
    ws6.cell(row=25, column=2).value = float(base_year_data['Total Debt'])
    ws6.cell(row=26, column=2).value = float(base_year_data['Minority Interest'])
    ws6.cell(row=28, column=2).value = base_year_data['Outstanding Shares']
    ws6.cell(row=33, column=2).value = float(base_year_data['Invested Capital'])

    # Apply number format to historical summary data (ws2)
    AMOUNT_ROWS = {'Revenue', 'EBIT', 'Depreciation & Amortization', 'Increase in Working Capital',
                   'Capital Expenditure', 'Total Reinvestments', 'Total Debt', 'Total Equity',
                   'Minority Interest', 'Cash & Cash Equivalents', 'Total Investments', 'Invested Capital'}
    RATIO_ROWS = {'Revenue Growth', 'EBIT Growth', 'EBIT Margin', 'Tax Rate', 'Revenue to Invested Capital',
                  'Debt to Assets', 'Cost of Debt', 'ROIC', 'ROE', 'Dividend Yield', 'Payout Ratio'}
    for row in ws2.iter_rows(min_row=2):
        row_label = row[0].value
        if row_label in AMOUNT_ROWS:
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        elif row_label in RATIO_ROWS:
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

    # Write gap analysis to a new sheet if available
    if gap_analysis_result and gap_analysis_result.get('analysis_text'):
        ws_gap = wb.create_sheet('Gap Analysis')
        currency = gap_analysis_result.get('currency', '')
        ws_gap.cell(row=1, column=1).value = 'DCF 估值 vs 当前股价 差异分析'
        from openpyxl.styles import Font
        ws_gap.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_gap.cell(row=3, column=1).value = f"当前股价: {gap_analysis_result['current_price']:.2f} {currency}"
        ws_gap.cell(row=4, column=1).value = f"DCF 估值: {gap_analysis_result['dcf_price']:.2f} {currency}"
        ws_gap.cell(row=5, column=1).value = f"差异: {gap_analysis_result['gap_pct']:+.1f}%"
        if gap_analysis_result.get('adjusted_price') is not None:
            ws_gap.cell(row=6, column=1).value = f"修正后估值: {gap_analysis_result['adjusted_price']:,.2f} {currency}"

        # Write analysis text line by line starting from row 8
        import re as _re
        analysis_text = _re.sub(r'\n?\s*ADJUSTED_PRICE:.*$', '', gap_analysis_result['analysis_text']).strip()
        for i, line in enumerate(analysis_text.split('\n'), start=8):
            ws_gap.cell(row=i, column=1).value = line
        ws_gap.column_dimensions['A'].width = 100

    for ws in [ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)

def main(args):
    while True:
        print("\nPlease enter the stock symbol and the financial data period to continue...\n")
        ticker = input('Enter the stock symbol (e.g., AAPL): ')
        args.t = ticker

        period = input('Select the financial data period (annual/quarter): ')
        args.period = period

        if period == 'annual':
            HISTORICAL_DATA_PERIODS = HISTORICAL_DATA_PERIODS_ANNUAL
        elif period == 'quarter':
            HISTORICAL_DATA_PERIODS = HISTORICAL_DATA_PERIODS_QUARTER
        else:
            raise ValueError("Invalid period. Please enter 'annual' or 'quarter'.")

        financial_data = get_historical_financials(args.t, args.period, args.apikey, HISTORICAL_DATA_PERIODS)
        if financial_data is None:
            print("Error: Failed to fetch financial data. Please check your API key and ticker symbol.")
            continue
        summary_df = financial_data['summary']
        company_info = get_company_share_float(args.t, args.apikey)
        company_profile = fetch_company_profile(args.t, args.apikey)
        company_name = company_profile.get('companyName', 'N/A')

        base_year_col = summary_df.columns[0]
        base_year_data = summary_df[base_year_col].copy()

        print(f"\n{company_name} Historical Financial Data (Summary, in millions):")
        formatted_summary_df = format_summary_df(summary_df)
        print(formatted_summary_df.to_string())

        if period == 'quarter':
            print("\nWarning: Valuation requires annual financial data. Please switch to 'annual' period to proceed.")
            exit_program = input('\nExit program? (y/n): ').strip().lower()
            if exit_program.lower() == 'y':
                print("Exiting...")
                break
            else:
                continue

        cont = input('\nProceed with valuation? (y/n): ').strip().lower()
        if cont.lower() != 'y':
            exit_program = input('Exit program? (y/n): ').strip().lower()
            if exit_program.lower() == 'y':
                print("Exiting...")
                break
            else:
                continue

        try:
            calendar_year = summary_df.index[summary_df.index.get_loc('Calendar Year')]
            base_year = int(calendar_year)
        except KeyError:
            base_year = int(summary_df.columns[0])

        outstanding_shares = company_info.get('outstandingShares', 0)
        base_year_data['Outstanding Shares'] = outstanding_shares
        base_year_data['Average Tax Rate'] = financial_data['average_tax_rate']
        base_year_data['Revenue Growth'] = financial_data['summary'].loc['Revenue Growth', base_year_col]
        base_year_data['Total Reinvestments'] = financial_data['summary'].loc['Total Reinvestments', base_year_col]

        print(f"\nThe base year used for cashflow forecast is {base_year}.")

        # Calculate WACC silently (details shown later during parameter review)
        average_tax_rate = base_year_data['Average Tax Rate']
        wacc, total_equity_risk_premium, wacc_details = calculate_wacc(base_year_data, company_profile, args.apikey, verbose=False)

        # Determine mode: AI or manual
        use_ai = not args.manual
        ai_params = None

        if use_ai:
            try:
                ai_result = analyze_company(
                    ticker=ticker,
                    summary_df=summary_df,
                    base_year_data=base_year_data,
                    company_profile=company_profile,
                    calculated_wacc=wacc,
                    calculated_tax_rate=average_tax_rate,
                    base_year=base_year,
                )
                ai_params = interactive_review(ai_result, wacc, average_tax_rate, company_profile, wacc_details)
            except Exception as e:
                print(f"\nAI 分析出错: {e}")
                print("自动回退到手工输入模式...\n")

        if ai_params is not None:
            # AI mode succeeded — build valuation_params from reviewed results
            ronic_match = ai_params.pop("ronic_match_wacc", True)
            if ronic_match:
                ronic = get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM
            else:
                ronic = get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM + TERMINAL_RONIC_PREMIUM

            valuation_params = {
                'base_year': base_year,
                'revenue_growth_1': ai_params['revenue_growth_1'],
                'revenue_growth_2': ai_params['revenue_growth_2'],
                'ebit_margin': ai_params['ebit_margin'],
                'convergence': ai_params['convergence'],
                'revenue_invested_capital_ratio_1': ai_params['revenue_invested_capital_ratio_1'],
                'revenue_invested_capital_ratio_2': ai_params['revenue_invested_capital_ratio_2'],
                'revenue_invested_capital_ratio_3': ai_params['revenue_invested_capital_ratio_3'],
                'tax_rate': ai_params['tax_rate'],
                'wacc': ai_params['wacc'],
                'terminal_wacc': get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM,
                'ronic': ronic,
                'risk_free_rate': get_risk_free_rate(company_profile.get('country', 'United States'))
            }
        else:
            # Manual mode
            print("\nEnter the following inputs...\n")
            revenue_growth_1 = float(input('Enter the annual revenue growth rate for Year 1 (%): '))
            revenue_growth_2 = float(input('Enter the Compound annual revenue growth rate for Years 2-5 (%): '))
            ebit_margin = float(input('Enter the target EBIT margin (%): '))
            convergence = float(input('Enter the number of years to reach the target EBIT margin: '))
            revenue_invested_capital_ratio_1 = float(input('Enter the revenue to invested capital ratio for Year 1: '))
            revenue_invested_capital_ratio_2 = float(input('Enter the revenue to invested capital ratio for Years 3-5: '))
            revenue_invested_capital_ratio_3 = float(input('Enter the revenue to invested capital ratio for Years 5-10: '))

            tax_rate_input = input(f"\nCalculated Average Tax Rate: {average_tax_rate:.1%}. Press Enter to accept as tax rate or enter a new value (e.g., 25 for 25%): ")
            if tax_rate_input.strip() == "":
                tax_rate = average_tax_rate * 100
            else:
                tax_rate = float(tax_rate_input)

            print_wacc_details(wacc_details)
            wacc_input = input(f"\nCalculated WACC: {wacc:.1%}. Press Enter to accept as discount rate or enter a new value (e.g., 8 for 8%): ")
            if wacc_input.strip() == "":
                wacc = wacc * 100
            else:
                wacc = float(wacc_input)

            cont = input('Will ROIC match terminal WACC beyond year 10? (y/n): ').strip().lower()
            if cont.lower() == 'y':
                ronic = get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM
            else:
                ronic = get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM + TERMINAL_RONIC_PREMIUM

            valuation_params = {
                'base_year': base_year,
                'revenue_growth_1': revenue_growth_1,
                'revenue_growth_2': revenue_growth_2,
                'ebit_margin': ebit_margin,
                'convergence': convergence,
                'revenue_invested_capital_ratio_1': revenue_invested_capital_ratio_1,
                'revenue_invested_capital_ratio_2': revenue_invested_capital_ratio_2,
                'revenue_invested_capital_ratio_3': revenue_invested_capital_ratio_3,
                'tax_rate': tax_rate,
                'wacc': wacc,
                'terminal_wacc': get_risk_free_rate(company_profile.get('country', 'United States')) + TERMINAL_RISK_PREMIUM,
                'ronic': ronic,
                'risk_free_rate': get_risk_free_rate(company_profile.get('country', 'United States'))
            }

        results = calculate_dcf(base_year_data, valuation_params, financial_data, company_info, company_profile)
        print_dcf_results(results, company_name)

        print("\nRunning sensitivity analysis...")
        sensitivity_table = sensitivity_analysis(base_year_data, valuation_params, financial_data, company_info, company_profile)
        print("\nSensitivity Analysis - Revenue Growth vs EBIT Margin (Price per Share):")
        print(sensitivity_table)

        print("\nRunning WACC sensitivity analysis...")
        wacc_table = wacc_sensitivity_analysis(base_year_data, valuation_params, financial_data, company_info, company_profile)
        print("\nSensitivity Analysis - WACC (Price per Share):")
        print(wacc_table.T.to_string())

        # AI gap analysis: compare DCF valuation vs current stock price
        gap_analysis_result = None
        if not args.manual:
            try:
                gap_analysis_result = analyze_valuation_gap(ticker, company_profile, results, valuation_params, summary_df, base_year)
            except Exception as e:
                print(f"\n估值差异分析出错: {e}")

        export_to_excel = input("\nDo you want to export the valuation results to Excel? (y/n): ").strip().lower()
        if export_to_excel == 'y':
            filename = os.path.join(EXCEL_OUTPUT_DIR, f"{company_name}_valuation_{date.today().strftime('%Y%m%d')}.xlsx")
            write_to_excel(filename, base_year_data, financial_data, valuation_params, company_profile, total_equity_risk_premium, gap_analysis_result)
            print(f"\nValuation results saved to {filename}")
        else:
            print("\nSkipping Excel export.")

        cont = input("\nValuation completed. Exit program? (y/n): ").strip().lower()
        if cont.lower() == 'y':
            print("Exiting...")
            break

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--apikey', help='API key for financialmodelingprep.com', default=os.environ.get('FMP_API_KEY'))
    parser.add_argument('--manual', action='store_true', help='Force manual input mode (skip AI analysis)')
    args = parser.parse_args()
    main(args)